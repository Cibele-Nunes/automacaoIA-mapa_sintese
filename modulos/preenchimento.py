from config import *
import pandas as pd
from shutil import copyfile
from openpyxl import load_workbook

# ==========================================================
# CARREGAR CSV OFICIAL
# ==========================================================

def carregar_csv_oficial(ano, mes):

    ano_mes = f"{ano}_{mes}"
    caminho = CSV_OFICIAL / f"{ano_mes}_dataframe_oficial.csv"

    if not caminho.exists():
        raise FileNotFoundError(f"CSV oficial não encontrado: {caminho}")

    df = pd.read_csv(caminho, sep=";", encoding="utf-8-sig")

    print("✔ CSV oficial carregado:", caminho.name)

    return df


# ==========================================================
# LIMPEZA E PADRONIZAÇÃO
# ==========================================================

def preparar_dados(df):

    df = df.copy()

    # Corrigir NOTA
    df["nota"] = (
        df["nota"]
        .astype(str)
        .str.replace(",", ".", regex=False)
        .str.strip()
    )

    df["nota"] = pd.to_numeric(df["nota"], errors="coerce")

    # Corrigir DATA (FORMATO BRASILEIRO)
    df["data"] = pd.to_datetime(df["data"], dayfirst=True, errors="coerce")

    # presença
    df["presenca"] = df["nota"].apply(
        lambda x: "AUSENTE" if pd.isna(x) else "PRESENTE"
    )

    # resultado oficial
    df["resultado"] = df.apply(
        lambda r: "REPROVADO" if r["presenca"]=="AUSENTE"
        else ("APROVADO" if r["nota"]>=5 else "REPROVADO"),
        axis=1
    )

    return df

def remover_registros_invalidos(df):
    df = df.copy()

    # normalizar vazios
    for col in ["area", "etapa", "nota"]:
        df[col] = df[col].astype(str).str.strip()
        df.loc[df[col] == "", col] = None
        df.loc[df[col] == "nan", col] = None

    # regra: não possui informação acadêmica nenhuma
    invalido = (
        df["area"].isna() &
        df["etapa"].isna() &
        df["nota"].isna()
    )

    removidos = df[invalido]["nome"].unique()
    df_limpo = df[~invalido]

    print("Registros removidos automaticamente (sem dados acadêmicos):")
    for n in removidos:
        print(" -", n)

    print("\nTotal removido:", len(df) - len(df_limpo))

    return df_limpo

def validar_csv_oficial(df):
    """
    Valida o CSV revisado antes de preencher o Excel.
    BLOCO ANTI-ERRO HUMANO.
    """

    # 🔧 limpeza inicial
    df = remover_registros_invalidos(df)

    print("Após limpeza:", len(df))

    erros = []

    # garantir tipo numérico
    df["nota"] = pd.to_numeric(df["nota"], errors="coerce")

    # 1) Nota em aluno ausente
    erro_nota_ausente = df[
        (df["presenca"] == "AUSENTE") &
        (df["nota"].notna())
    ]
    if len(erro_nota_ausente) > 0:
        erros.append(f"Há {len(erro_nota_ausente)} alunos AUSENTES com nota.")

    # 2) Resultado inválido
    valores_validos = ["APROVADO", "REPROVADO"]
    erro_resultado = df[
        ~df["resultado"].isin(valores_validos)
    ]
    if len(erro_resultado) > 0:
        erros.append("Existem valores inválidos na coluna RESULTADO.")

    # 3) Área vazia
    erro_area = df[
        df["area"].isna() |
        (df["area"].astype(str).str.strip() == "")
    ]
    if len(erro_area) > 0:
        erros.append("Existem alunos sem área.")

    # 4) Nome vazio
    erro_nome = df[
        df["nome"].isna() |
        (df["nome"].astype(str).str.strip() == "")
    ]
    if len(erro_nome) > 0:
        erros.append("Existem alunos sem nome.")

    # 5) Nota incompatível com resultado
    erro_resultado_nota = df[
        (df["nota"] >= 5.0) &
        (df["resultado"] == "REPROVADO")
    ]
    if len(erro_resultado_nota) > 0:
        erros.append("Existem alunos APROVADOS por nota mas marcados como REPROVADO.")

    erro_resultado_nota2 = df[
        (df["nota"] < 5.0) &
        (df["presenca"] == "PRESENTE") &
        (df["resultado"] == "APROVADO")
    ]
    if len(erro_resultado_nota2) > 0:
        erros.append("Existem alunos com nota abaixo de 5,0 marcados como APROVADO.")

    # 6) Nota fora do intervalo
    erro_intervalo = df[
        (df["nota"] < 0) |
        (df["nota"] > 10)
    ]
    if len(erro_intervalo) > 0:
        erros.append("Existem notas fora do intervalo 0 a 10.")

    # resultado
    if erros:
        print("ERROS ENCONTRADOS:")
        for e in erros:
            print("- ", e)
    else:
        print("CSV validado com sucesso — pronto para continuar o pipeline")

    print("Total de registros:", len(df))

    return df

# =========================================================
# CARREGAMENTO E ABERTURA DO ARQUIVO ANUAL
# =========================================================

def carregar_arquivo_anual(ano):
    """
    Carrega (ou cria) o arquivo anual do mapa síntese.
    """

    ARQUIVO_ANUAL = PASTA_MAPA_ANUAL / f"MAPA_SINTESE_{ano}.xlsx"

    # cria se não existir
    if not ARQUIVO_ANUAL.exists():
        print("Arquivo anual não encontrado.")
        print("Criando a partir do modelo base...")
        copyfile(ARQUIVO_MODELO, ARQUIVO_ANUAL)
    else:
        print("Arquivo anual encontrado. Será atualizado.")

    # carrega workbook
    workbook = load_workbook(ARQUIVO_ANUAL)

    ws_fundamental = workbook["ENSINO FUNDAMENTAL"]
    ws_medio = workbook["ENSINO MÉDIO"]

    # 🔍 DEBUG (mantido, como você já usa)
    print("\n--- DEBUG EXCEL ---")
    print("Arquivo anual:", ARQUIVO_ANUAL)
    print("Existe?", ARQUIVO_ANUAL.exists())
    print("Tamanho:", ARQUIVO_ANUAL.stat().st_size)

    print("Modelo:", ARQUIVO_MODELO)
    print("Modelo existe?", ARQUIVO_MODELO.exists())

    print("Sheets disponíveis:")
    print(workbook.sheetnames)
    print("-------------------\n")

    return workbook, ws_fundamental, ws_medio, ARQUIVO_ANUAL

def padronizacao_final(df):
    """
    Padronização final antes de escrever no Excel.
    Garante consistência total dos dados.
    """

    df_final = df.copy()

    # padronização de colunas
    df_final.columns = df_final.columns.str.lower().str.strip()

    # padronização de conteúdo
    df_final["nome"] = df_final["nome"].astype(str).str.strip()
    df_final["area"] = df_final["area"].astype(str).str.upper().str.strip()
    df_final["etapa"] = df_final["etapa"].astype(str).str.upper().str.strip()
    df_final["presenca"] = df_final["presenca"].astype(str).str.upper().str.strip()
    df_final["resultado"] = df_final["resultado"].astype(str).str.upper().str.strip()

    df_final["nota"] = pd.to_numeric(df_final["nota"], errors="coerce")

    print("Dados oficiais carregados para escrita:", len(df_final))

    return df_final

def aplicar_mapa_areas(df):
    """
    Traduz as áreas do CSV para o padrão do Excel.
    """

    df = df.copy()

    df["area_estrutura"] = df["area"].map(MAPA_ESTRUTURA_EXCEL)

    # validação obrigatória
    nao_traduzidas = df[df["area_estrutura"].isna()][["area"]].drop_duplicates()

    if len(nao_traduzidas) > 0:
        print("⚠️ Áreas não reconhecidas:")
        print(nao_traduzidas)
    else:
        print("Todas as áreas reconhecidas pelo modelo!")

    return df

# ==========================================
# EXTRAIR MÊS DA DATA
# ==========================================

def extrair_mes(df):
    """
    Cria a coluna 'mes' a partir da coluna 'data'.
    """

    df = df.copy()

    df["mes"] = df["data"].dt.month.map(MESES_PT)

    print("Meses encontrados:")
    print(df["mes"].unique())

    return df

def validar_mes_unico(df):
    """
    Garante que o CSV contém apenas um mês.
    """

    meses_csv = df["mes"].dropna().unique()

    if len(meses_csv) != 1:
        raise ValueError(
            f"CSV deve conter apenas 1 mês. Encontrado: {meses_csv}"
        )

    mes_atual = meses_csv[0]

    print("Mês detectado no CSV:", mes_atual)

    return mes_atual

# diagnóstico de debug - usar só quando precisar
# def diagnostico_datas(df):
#    print("Tipo da coluna data:", df["data"].dtype)
#    print("Meses numéricos:", df["data"].dt.month.unique())

# ============================================================
# CRIA TABELA RESUMO OFICIAL PARA PREENCHER O MODELO
# ============================================================

def gerar_resumo(df):
    """
    Gera a tabela resumo para preenchimento do Excel.
    """

    df = df.copy()

    # PADRONIZAÇÃO DE SEGURANÇA (ANTES DO GROUPBY)
    df["presenca"] = df["presenca"].astype(str).str.upper().str.strip()
    df["resultado"] = df["resultado"].astype(str).str.upper().str.strip()
    df["turno"] = df["turno"].astype(str).str.upper().str.strip()
    df["etapa"] = df["etapa"].astype(str).str.upper().str.strip()
    df["area_estrutura"] = df["area_estrutura"].astype(str).str.upper().str.strip()
    df["mes"] = df["mes"].astype(str).str.upper().str.strip()

    # indicadores
    df["INSCRITOS"] = 1
    df["PRESENTES"] = (df["presenca"] == "PRESENTE").astype(int)
    df["AUSENTES"] = (df["presenca"] == "AUSENTE").astype(int)
    df["APROVADOS"] = (df["resultado"] == "APROVADO").astype(int)
    df["REPROVADOS"] = (df["resultado"] == "REPROVADO").astype(int)

    # agrupamento
    resumo = (
        df.groupby(
            ["etapa", "mes", "area_estrutura", "turno"],
            as_index=False
        )
        .agg({
            "INSCRITOS": "sum",
            "PRESENTES": "sum",
            "AUSENTES": "sum",
            "APROVADOS": "sum",
            "REPROVADOS": "sum"
        })
    )

    print("Resumo gerado com sucesso.")
    print("Total de linhas no resumo:", len(resumo))

    return resumo


def escrever_no_modelo(ws, etapa_nome, resumo):
    """
    Preenche UMA planilha inteira (fundamental ou médio)
    usando a tabela resumo já calculada
    """

    df_etapa = resumo[resumo["etapa"] == etapa_nome]

    for _, linha in df_etapa.iterrows():

        mes = linha["mes"]
        area = linha["area_estrutura"]
        turno = linha["turno"]

        if mes not in LINHAS_MESES:
            print("Mês ignorado:", mes)
            continue

        if area not in OFFSET_AREAS:
            print("Área ignorada:", area)
            continue

        # calcula linha do excel
        linha_excel = LINHAS_MESES[mes] + OFFSET_AREAS[area]

        # escreve cada campo
        for campo in ["INSCRITOS","PRESENTES","AUSENTES","APROVADOS","REPROVADOS"]:

            coluna = COLUNAS_TURNO[turno][campo]
            valor = int(linha[campo])

            ws[f"{coluna}{linha_excel}"] = valor

    print("Planilha preenchida:", etapa_nome)
    print(f"\n🔎 Etapa: {etapa_nome}")
    print("Linhas encontradas:", len(df_etapa))

def verificar_mes_preenchido(ws, mes):
    """
    Verifica se o mês já foi preenchido.
    Considera preenchido apenas se a célula
    NÃO contiver mais fórmula (ou seja, já foi escrita).
    """

    if mes not in LINHAS_MESES:
        return False

    linha_base = LINHAS_MESES[mes]

    for offset in OFFSET_AREAS.values():
        linha_excel = linha_base + offset

        for colunas_turno in COLUNAS_TURNO.values():
            for coluna in colunas_turno.values():

                cell = ws[f"{coluna}{linha_excel}"]

                # Se não for fórmula, significa que já foi escrita
                if cell.data_type != "f":
                    if cell.value not in (None, ""):
                        return True

    return False

def executar_preenchimento(ano, mes):

    df = carregar_csv_oficial(ano, mes)

    df = preparar_dados(df)

    df = validar_csv_oficial(df)

    df = padronizacao_final(df)

    df = aplicar_mapa_areas(df)

    df = extrair_mes(df)

    MES_ATUAL = validar_mes_unico(df)

    workbook, ws_fundamental, ws_medio, ARQUIVO_ANUAL = carregar_arquivo_anual(ano)

    ja_preenchido_fund = verificar_mes_preenchido(ws_fundamental, MES_ATUAL)
    ja_preenchido_medio = verificar_mes_preenchido(ws_medio, MES_ATUAL)

    if ja_preenchido_fund or ja_preenchido_medio:
        raise ValueError("Mês já preenchido")

    resumo = gerar_resumo(df)

    escrever_no_modelo(ws_fundamental, "ENSINO FUNDAMENTAL", resumo)
    escrever_no_modelo(ws_medio, "ENSINO MÉDIO", resumo)

    workbook.save(ARQUIVO_ANUAL)

    # caso dezembro
    if MES_ATUAL == "DEZEMBRO":

        ARQUIVO_FINAL = (
            PASTA_RESULTADOS
            / "arquivo_final"
            / f"SALVADOR - ROBERTO SANTOS - MAPA SINTESE POR UC MARÇO A DEZEMBRO {ANO}.xlsx"
    )

        ARQUIVO_FINAL.parent.mkdir(parents=True, exist_ok=True)

        workbook.save(ARQUIVO_FINAL)

        print("Arquivo final gerado com sucesso!")
        print("Local:", ARQUIVO_FINAL)

